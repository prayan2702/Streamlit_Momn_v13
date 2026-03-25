"""
momn_streamlit_app_v12.py
=========================
Momentum Screener + Portfolio Rebalancer — v12

Changes vs v11:
  • Universe dropdown (Nifty50/100/200/250/500/N750/AllNSE)
    – AllNSE: EQUITY_L.csv upload + GitHub fallback
    – Others : auto-fetch from GitHub (same URLs as v10/main_window.py)
  • data_service.fetch_data() integration (YFinance / Upstox / Angel One)
  • Upstox & Angel One auth shown in sidebar (upstox_auth / angelone_auth)
  • Excel export — v10 formatting: 4 sheets, color highlights, rank summary
  • Step 3 rebalance logic = v10 (portfolio vs top-N, reasons for exit)
    + Quick Rebalance panel (capital add, brokerage, per-stock orders)
  • Rebalance Sheet quick-link → Apps Script URL
"""

import io
import time
import datetime
import warnings

import numpy as np
import pandas as pd
import streamlit as st
import requests

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

warnings.filterwarnings("ignore")

# ── Local modules ──────────────────────────────────────────────
try:
    from calculations import build_dfStats, apply_filters
    _CALCS_AVAILABLE = True
except ImportError:
    _CALCS_AVAILABLE = False

import yfinance as yf  # always available (in requirements.txt)

# ── data_service (may fail if SmartApi/pyotp not installed) ───
_DS_AVAILABLE   = False
_DS_IMPORT_ERR  = ""
try:
    from data_service import fetch_data
    _DS_AVAILABLE = True
except Exception as _e:
    _DS_IMPORT_ERR = str(_e)

# ── upstox_auth ───────────────────────────────────────────────
_UPSTOX_AVAILABLE = False
try:
    from upstox_auth import get_upstox_access_token
    _UPSTOX_AVAILABLE = True
except Exception:
    pass

# ── angelone_auth ─────────────────────────────────────────────
_ANGEL_AVAILABLE = False
try:
    from angelone_auth import get_angelone_client
    _ANGEL_AVAILABLE = True
except Exception:
    pass

# ── Inline YFinance fetcher (fallback when data_service fails) ─
def _fetch_yfinance_inline(symbols_ns, start_date, end_date,
                            progress_bar, status_text, chunk_size=15):
    """Pure yfinance fetch — no data_service dependency."""
    close_chunks, high_chunks, vol_chunks = [], [], []
    failed = []
    total  = len(symbols_ns)
    for k in range(0, total, chunk_size):
        chunk = symbols_ns[k:k + chunk_size]
        pct   = min((k + chunk_size) / total, 1.0)
        status_text.markdown(f"⏳ **Fetching {k+1}–{min(k+chunk_size, total)} / {total}**")
        progress_bar.progress(pct * 0.88)
        try:
            raw = yf.download(chunk, start=start_date, end=end_date,
                              progress=False, auto_adjust=True, threads=True,
                              multi_level_index=False)
            if not raw.empty:
                close_chunks.append(raw["Close"])
                high_chunks.append(raw["High"])
                vol_val = raw["Close"].multiply(raw.get("Volume", 1))
                vol_chunks.append(vol_val)
        except Exception as e:
            failed.extend(chunk)
        time.sleep(0.5)

    if not close_chunks:
        return None, None, None, failed

    close  = pd.concat(close_chunks,  axis=1)
    high   = pd.concat(high_chunks,   axis=1)
    volume = pd.concat(vol_chunks,    axis=1)
    close  = close.loc[:,  ~close.columns.duplicated()]
    high   = high.loc[:,   ~high.columns.duplicated()]
    volume = volume.loc[:, ~volume.columns.duplicated()]
    return close, high, volume, failed

# ═══════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Momn Screener v12",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.app-header {
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 50%, #0f172a 100%);
    border-bottom: 2px solid #334155;
    padding: 16px 24px; margin: -1rem -1rem 1.5rem -1rem;
    display:flex; align-items:center; justify-content:space-between;
}
.app-title { color:#f8fafc; font-size:20px; font-weight:700; }
.app-title span { color:#38bdf8; }
.app-subtitle { color:#94a3b8; font-size:12px; margin-top:2px; }

.step-bar {
    display:flex; align-items:center; gap:0;
    background:#f8fafc; border:1px solid #e2e8f0;
    border-radius:12px; padding:10px 18px;
    margin-bottom:1.2rem; overflow-x:auto;
}
.step-item {
    display:flex; align-items:center; gap:8px;
    padding:7px 14px; border-radius:8px;
    font-size:13px; font-weight:600; white-space:nowrap;
}
.step-item.done   { background:#dcfce7; color:#15803d; }
.step-item.active { background:#dbeafe; color:#1d4ed8; box-shadow:0 0 0 2px #3b82f6; }
.step-item.pending{ color:#94a3b8; }
.step-circle {
    width:24px; height:24px; border-radius:50%;
    display:flex; align-items:center; justify-content:center;
    font-size:11px; font-weight:700; flex-shrink:0;
}
.done .step-circle   { background:#16a34a; color:#fff; }
.active .step-circle { background:#2563eb; color:#fff; }
.pending .step-circle{ background:#e2e8f0; color:#94a3b8; }
.step-connector { width:28px; height:2px; background:#e2e8f0; flex-shrink:0; }
.step-connector.done-line { background:#16a34a; }

.metric-row  { display:flex; gap:10px; flex-wrap:wrap; margin:10px 0; }
.metric-card {
    background:#f8fafc; border:1px solid #e2e8f0;
    border-radius:10px; padding:10px 16px; min-width:130px;
}
.metric-label { font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:.4px; font-weight:600; }
.metric-value { font-size:21px; font-weight:700; color:#0f172a; margin-top:2px; }
.metric-value.green { color:#16a34a; }
.metric-value.red   { color:#dc2626; }
.metric-value.blue  { color:#2563eb; }

.section-hdr {
    font-size:14px; font-weight:700; color:#0f172a;
    border-bottom:2px solid #3b82f6; padding-bottom:5px;
    margin:1.1rem 0 .7rem;
}
.nse-link-box {
    background:#eff6ff; border:1px solid #bfdbfe; border-radius:10px;
    padding:12px 16px; display:flex; align-items:center; gap:10px; margin:10px 0;
}
.nse-link-box a { color:#2563eb; font-weight:600; font-size:13px; }
.nse-link-box .hint { font-size:11px; color:#64748b; margin-top:2px; }
.chip { display:inline-block; padding:3px 9px; border-radius:16px; font-size:11px; font-weight:700; margin:2px; }
.chip-sell { background:#fee2e2; color:#dc2626; }
.chip-buy  { background:#dcfce7; color:#16a34a; }
.chip-hold { background:#f1f5f9; color:#475569; }
.reb-strip {
    display:flex; gap:12px; flex-wrap:wrap;
    background:#f8fafc; border:1px solid #e2e8f0;
    border-radius:10px; padding:10px 16px; margin:8px 0;
}
.reb-stat .label { font-size:9px; color:#64748b; text-transform:uppercase; font-weight:600; }
.reb-stat .val   { font-size:16px; font-weight:700; }
.reb-stat .val.r { color:#dc2626; }
.reb-stat .val.g { color:#16a34a; }
.reb-stat .val.b { color:#2563eb; }
.reb-stat .val.p { color:#7c3aed; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════
UNIVERSES    = ['Nifty50','Nifty100','Nifty200','Nifty250','Nifty500','N750','AllNSE']
API_OPTIONS  = ["YFinance","Upstox","Angel One"]
RANKING_MAP  = {
    "AvgZScore 12M/6M/3M":    "avgZScore12_6_3",
    "AvgZScore 12M/9M/6M/3M": "avgZScore12_9_6_3",
    "AvgSharpe 12M/6M/3M":    "avgSharpe12_6_3",
    "AvgSharpe 9M/6M/3M":     "avgSharpe9_6_3",
    "AvgSharpe 12M/9M/6M/3M": "avg_All",
    "Sharpe12M":               "sharpe12M",
    "Sharpe3M":                "sharpe3M",
}
PORTFOLIO_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vS4HDgiell4n1kd08OnlzOQobfPzeDtVyWJ8gETFlYbz27qhOmfqKZOoIXZItRQEq5ANATYIcZJm0gk"
    "/pub?output=csv"
)
APPS_SCRIPT_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbxIUgGxTf-zNJEq-bCMDMXYYCpZd9FeW4-06ovt1skKXmVjVcnIMv_GCmQ85DMH6xev/exec"
)
GITHUB_BASE = "https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main"

USERNAME = "prayan"
PASSWORD = "prayan"

# ═══════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ═══════════════════════════════════════════════════════════════
_defaults = {
    "logged_in":      False,
    "current_step":   1,
    "universe":       "AllNSE",
    "symbols":        None,       # list[str] — .NS suffixed
    "eq_df":          None,       # only for AllNSE CSV upload
    "dfStats":        None,
    "dfFiltered":     None,
    "failed_blank":   [],
    "reb_portfolio":  None,
    "sell_list":      None,
    "buy_list":       None,
    "rebalance_table": None,
    "lookback_date":  datetime.date.today(),
    "ranking_method": "avgZScore12_6_3",
    "data_source":    "YFinance",
    "top_n_rank":     100,
    "screener_done":  False,
    "rebalance_done": False,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════
def fmt_inr(v):
    if pd.isna(v): return "—"
    v = int(round(v))
    if abs(v) >= 10_000_000: return f"₹{v/10_000_000:.1f}Cr"
    if abs(v) >= 100_000:    return f"₹{v/100_000:.1f}L"
    return f"₹{v:,}"

def step_html(current):
    steps = [(1,"Universe Setup"),(2,"Run Screener"),(3,"Plan Rebalance"),(4,"Apply & Export")]
    html = '<div class="step-bar">'
    for i,(n,label) in enumerate(steps):
        cls = "done" if n < current else ("active" if n == current else "pending")
        sym = "✓"    if n < current else str(n)
        html += f'<div class="step-item {cls}"><div class="step-circle">{sym}</div>{label}</div>'
        if i < len(steps)-1:
            lc = "done-line" if n < current else ""
            html += f'<div class="step-connector {lc}"></div>'
    return html + '</div>'

def metric_card(label, value, color=""):
    return f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value {color}">{value}</div></div>'

def parse_equity_csv(f) -> pd.DataFrame:
    df = pd.read_csv(f, skipinitialspace=True)
    df.columns = [c.strip() for c in df.columns]
    if 'SERIES' in df.columns:
        df = df[df['SERIES'].str.strip() == 'EQ'].copy()
    df['SYMBOL'] = df['SYMBOL'].str.strip().str.upper()
    return df.reset_index(drop=True)

def load_symbols_from_github(universe: str) -> list:
    """Returns list of .NS symbols for the chosen universe (not AllNSE)."""
    if universe == 'N750':
        url = f"{GITHUB_BASE}/ind_niftytotalmarket_list.csv"
    else:
        url = f"{GITHUB_BASE}/ind_{universe.lower()}list.csv"
    df = pd.read_csv(url)
    df['Yahoo_Symbol'] = df['Symbol'].astype(str).str.strip() + '.NS'
    return df['Yahoo_Symbol'].tolist()

def build_dates(end_date: datetime.date) -> dict:
    from dateutil.relativedelta import relativedelta
    end = datetime.datetime.combine(end_date, datetime.time())
    return {
        'startDate': datetime.datetime(2000, 1, 1),
        'endDate':   end,
        'date12M':   end - relativedelta(months=12),
        'date9M':    end - relativedelta(months=9),
        'date6M':    end - relativedelta(months=6),
        'date3M':    end - relativedelta(months=3),
        'date1M':    end - relativedelta(months=1),
    }

# ── v10-identical Excel formatting ────────────────────────────
def format_excel_unfiltered(file_name, universe, top_n):
    """Format 'Unfiltered Stocks' sheet — exact v10 logic."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = 'A2'
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    no_fill   = PatternFill(start_color="d6b4fc", end_color="d6b4fc", fill_type="solid")
    bold_font = Font(bold=True)
    headers   = [c.value for c in ws[1]]

    def ci(name): return headers.index(name) + 1 if name in headers else None

    rank_threshold = top_n
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    idx = {k: ci(k) for k in ['volm_cr','Close','dma200d','AWAY_ATH','roc12M',
                                'circuit','roc1M','circuit5','Ticker','Rank']}

    for row in range(2, ws.max_row + 1):
        failed = False
        def v(col): return ws.cell(row=row, column=col).value if col else None
        def mark(col):
            nonlocal failed
            ws.cell(row=row, column=col).fill = no_fill
            ws.cell(row=row, column=col).font = bold_font
            failed = True
        if (vol := v(idx['volm_cr']))  is not None and vol < 1:           mark(idx['volm_cr'])
        cl = v(idx['Close']); dm = v(idx['dma200d'])
        if cl is not None and dm is not None and cl <= dm:                 mark(idx['Close'])
        if (aa := v(idx['AWAY_ATH']))  is not None and aa <= -25:         mark(idx['AWAY_ATH'])
        roc = v(idx['roc12M'])
        if roc is not None and roc <= 5.5:                                 mark(idx['roc12M'])
        if (ci_ := v(idx['circuit']))  is not None and ci_ >= 20:         mark(idx['circuit'])
        if cl is not None and cl <= 30:                                    mark(idx['Close'])
        if (c5 := v(idx['circuit5']))  is not None and c5 > 10:           mark(idx['circuit5'])
        if roc is not None and roc > 1000:                                 mark(idx['roc12M'])
        if failed and idx['Ticker']:
            ws.cell(row=row, column=idx['Ticker']).fill = no_fill
        if idx['Rank'] and (rk := v(idx['Rank'])) is not None and rk <= rank_threshold:
            ws.cell(row=row, column=idx['Rank']).fill = green_fill

    # ATH round
    ath_col = ci('ATH')
    if ath_col:
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=ath_col)
            if isinstance(c.value, (int, float)):
                c.value = round(c.value)
    wb.save(file_name)


def format_excel_filtered(file_name, universe, top_n):
    """Format 'Filtered Stocks' sheet — exact v10 logic."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Filtered Stocks"]
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = 'A2'
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # ATH round
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "ATH":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = round(c.value)
            break
    # AWAY_ATH % suffix
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "AWAY_ATH":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = f"{c.value}%"
            break

    # Rank highlight + summary
    rank_threshold = top_n
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Rank":
            rank_75_count = 0
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)) and c.value <= rank_threshold:
                    c.fill = green_fill
                    rank_75_count += 1
            total_filtered = ws.max_row - 1
            ws.append([])
            ws.append(["Summary"])
            summary_start = ws.max_row
            ws.append([f"Total Filtered Stocks: {total_filtered}"])
            ws.append([f"Number of Stocks within {rank_threshold} Rank: {rank_75_count}"])
            for r in ws.iter_rows(min_row=summary_start, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in r:
                    cell.font = Font(bold=True)
            break
    wb.save(file_name)


def format_simple_sheet(file_name, sheet_name):
    """Format 'Failed Downloads' and 'Portfolio Rebalancing' sheets — exact v10 logic."""
    wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        wb.save(file_name); return
    ws = wb[sheet_name]
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = 'A2'

    if sheet_name == "Portfolio Rebalancing":
        headers  = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sell_col = headers.index('Sell Stocks') + 1 if 'Sell Stocks' in headers else None
        buy_col  = headers.index('Buy Stocks')  + 1 if 'Buy Stocks'  in headers else None
        sell_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
        buy_fill  = PatternFill(start_color="D7FFD7", end_color="D7FFD7", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            if sell_col: ws.cell(row=r, column=sell_col).fill = sell_fill
            if buy_col:  ws.cell(row=r, column=buy_col).fill  = buy_fill

    if sheet_name == "Failed Downloads":
        headers   = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        stock_col = headers.index('Failed Stock') + 1 if 'Failed Stock' in headers else None
        org_fill  = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        if stock_col:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=stock_col).fill = org_fill
    wb.save(file_name)


# ═══════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════
def login_page():
    # Compact centered card — same visual size as v10
    st.markdown("""
    <style>
    /* Hide default streamlit padding on login page */
    .login-wrap { display:flex; justify-content:center; margin-top:80px; }
    .login-card {
        background:#fff; border:1px solid #e2e8f0; border-radius:12px;
        padding:36px 40px; width:360px; box-shadow:0 4px 16px rgba(0,0,0,0.08);
    }
    .login-title { font-size:22px; font-weight:700; color:#0f172a;
                   margin-bottom:24px; text-align:center; }
    </style>
    <div class="login-wrap"><div class="login-card">
      <div class="login-title">🔐 Login</div>
    </div></div>
    """, unsafe_allow_html=True)

    # Center the form using columns (mimics v10 narrow layout)
    _, mid, _ = st.columns([1, 1.2, 1])
    with mid:
        with st.form("login_form", clear_on_submit=True):
            u = st.text_input("Username")
            p = st.text_input("Password", type="password")
            if st.form_submit_button("Login", use_container_width=True):
                if u == USERNAME and p == PASSWORD:
                    st.session_state.logged_in = True
                    st.rerun()
                else:
                    st.error("Invalid username or password")

if not st.session_state.logged_in:
    login_page()
    st.stop()

# ═══════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
  <div>
    <div class="app-title">📈 <span>Momn</span> Screener + Rebalancer</div>
    <div class="app-subtitle">NSE Momentum Strategy — v12</div>
  </div>
  <div style="color:#94a3b8;font-size:12px;">prayan2702</div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Workflow Steps")
    step_labels = {1:"1️⃣ Universe Setup", 2:"2️⃣ Run Screener",
                   3:"3️⃣ Plan Rebalance", 4:"4️⃣ Apply & Export"}
    for s, lbl in step_labels.items():
        is_active = (st.session_state.current_step == s)
        is_done   = (s == 1 and st.session_state.symbols is not None) or \
                    (s == 2 and st.session_state.screener_done) or \
                    (s == 3 and st.session_state.rebalance_done)
        indicator = "✅" if is_done else ("▶" if is_active else "○")
        if st.button(f"{indicator} {lbl}", key=f"nav_{s}", use_container_width=True,
                     type="primary" if is_active else "secondary"):
            st.session_state.current_step = s; st.rerun()

    st.divider()
    st.markdown("### 🔧 Screener Settings")
    rm_display = st.selectbox("Ranking Method", list(RANKING_MAP.keys()), index=0)
    st.session_state.ranking_method = RANKING_MAP[rm_display]

    st.session_state.data_source = st.selectbox("Data Source", API_OPTIONS, index=0)
    st.session_state.lookback_date = st.date_input(
        "Lookback Date", value=st.session_state.lookback_date,
        max_value=datetime.date.today()
    )
    st.session_state.top_n_rank = st.number_input(
        "Top-N Rank", min_value=20, max_value=200, value=100, step=10
    )

    # ── API Authentication ──────────────────────────────────
    if st.session_state.data_source == "Upstox":
        st.divider()
        if _UPSTOX_AVAILABLE:
            get_upstox_access_token(sidebar=True)
        else:
            st.warning("⚠️ `pyotp` install nahi hai. YFinance fallback use hoga.")

    elif st.session_state.data_source == "Angel One":
        st.divider()
        if _ANGEL_AVAILABLE:
            get_angelone_client(sidebar=True)
        else:
            st.warning(
                "⚠️ Angel One ke liye `smartapi-python` + `pyotp` "
                "`requirements.txt` mein add karo. "
                "Abhi **YFinance** fallback use hoga."
            )

    st.divider()
    st.markdown("### 🔗 Quick Links")
    st.markdown(f"""
    <div style="font-size:12px;line-height:2.2;">
    <a href="https://www.nseindia.com/static/market-data/securities-available-for-trading" target="_blank">📥 DOWNLOAD NSE EQUITY_L.csv</a><br>
    <a href="{APPS_SCRIPT_URL}" target="_blank">⚖️ REBALANCER</a><br>
    <a href="https://docs.google.com/spreadsheets/d/1xb8xoW91HWeXBW8Zd99TobULSgwxcvfPaaYPlMLZmHI" target="_blank">📊 REBALANCE SHEET</a>
    </div>
    """, unsafe_allow_html=True)

    if st.button("🚪 Logout", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

# ── Step progress bar ──────────────────────────────────────────
st.markdown(step_html(st.session_state.current_step), unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# STEP 1 — UNIVERSE SETUP
# ═══════════════════════════════════════════════════════════════
if st.session_state.current_step == 1:
    st.markdown('<div class="section-hdr">Step 1 — Universe Setup</div>', unsafe_allow_html=True)

    c1, c2 = st.columns([1, 2])
    with c1:
        chosen_u = st.selectbox(
            "Select Universe",
            UNIVERSES,
            index=UNIVERSES.index(st.session_state.universe),
            help="AllNSE = NSE ki sabhi EQ stocks. Baaki = Nifty index lists (GitHub se auto-load)"
        )
        st.session_state.universe = chosen_u

    # ── AllNSE: CSV upload ────────────────────────────────────
    if chosen_u == "AllNSE":
        st.markdown("""
        <div class="nse-link-box">
          <div>📥</div>
          <div>
            <a href="https://www.nseindia.com/static/market-data/securities-available-for-trading" target="_blank">
              NSE — Securities Available for Trading</a>
            <div class="hint">EQUITY_L.csv download karo → browse karo neeche | Ya GitHub fallback automatically use hoga</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader("📂 Browse EQUITY_L.csv (optional — GitHub fallback available)",
                                     type=["csv"], key="equity_csv")
        if uploaded:
            try:
                eq_df = parse_equity_csv(uploaded)
                st.session_state.eq_df  = eq_df
                syms_ns = [s + ".NS" for s in eq_df["SYMBOL"].tolist()]
                st.session_state.symbols = syms_ns
                st.session_state.universe_label = f"AllNSE (CSV — {len(syms_ns):,} stocks)"
                st.success(f"✅ CSV loaded: **{len(syms_ns):,}** EQ stocks")
                st.dataframe(eq_df[["SYMBOL","NAME OF COMPANY"]].head(20),
                             use_container_width=True, height=280)
            except Exception as e:
                st.error(f"CSV parse error: {e}")

        if not st.session_state.symbols:
            st.info("💡 CSV upload nahi hua — GitHub fallback (NSE_EQ_ALL.csv) use hoga screener run pe.")
        else:
            n = len(st.session_state.symbols)
            st.markdown(f"""<div class="metric-row">
                {metric_card("Loaded Symbols", f"{n:,}", "green")}
                {metric_card("Universe", "AllNSE", "blue")}
            </div>""", unsafe_allow_html=True)

    # ── Other universes: auto-fetch info ─────────────────────
    else:
        st.info(f"📡 **{chosen_u}** ki symbol list screener run pe GitHub se auto-load hogi.\n\n"
                f"CSV upload ki zaroorat nahi hai.")
        # Pre-load symbols when user confirms
        if st.button("✅ Load Symbol List", type="primary"):
            with st.spinner(f"Loading {chosen_u} from GitHub…"):
                try:
                    syms_ns = load_symbols_from_github(chosen_u)
                    st.session_state.symbols = syms_ns
                    st.session_state.universe_label = f"{chosen_u} ({len(syms_ns)} stocks)"
                    st.success(f"✅ {chosen_u}: **{len(syms_ns)}** symbols loaded")
                except Exception as e:
                    st.error(f"Symbol load failed: {e}")

        if st.session_state.symbols and st.session_state.universe == chosen_u:
            n = len(st.session_state.symbols)
            st.markdown(f"""<div class="metric-row">
                {metric_card("Loaded Symbols", f"{n:,}", "green")}
                {metric_card("Universe", chosen_u, "blue")}
            </div>""", unsafe_allow_html=True)

    st.divider()
    # ── Next step button ──────────────────────────────────────
    if st.session_state.symbols or chosen_u != "AllNSE":
        if st.button("▶ Next: Run Screener →", type="primary"):
            if st.session_state.symbols is None and chosen_u != "AllNSE":
                # Will load during screener run
                pass
            st.session_state.current_step = 2; st.rerun()
    elif chosen_u == "AllNSE" and not st.session_state.symbols:
        if st.button("▶ Next: Run Screener → (GitHub fallback)", type="secondary"):
            st.session_state.current_step = 2; st.rerun()

# ═══════════════════════════════════════════════════════════════
# STEP 2 — RUN SCREENER
# ═══════════════════════════════════════════════════════════════
elif st.session_state.current_step == 2:
    st.markdown('<div class="section-hdr">Step 2 — Run Momentum Screener</div>', unsafe_allow_html=True)

    if not _CALCS_AVAILABLE:
        st.error("❌ `calculations.py` not found. Project folder mein rakh kar dobara run karo.")
        st.stop()
    if not _DS_AVAILABLE:
        st.warning(
            f"⚠️ `data_service.py` import failed (reason: `{_DS_IMPORT_ERR[:120]}`). "
            "**YFinance** inline fallback use hoga. "
            "Upstox/Angel One ke liye `pyotp` + `smartapi-python` ko `requirements.txt` mein add karo."
        )

    # ── Filter settings ───────────────────────────────────────
    with st.expander("🔧 Filter Settings", expanded=True):
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            use_dma200 = st.checkbox("Close > 200-day DMA",  value=True)
            use_roc12  = st.checkbox("12M ROC > 5.5%",        value=True)
            use_roc_cap= st.checkbox("12M return < 1000x",    value=True)
        with fc2:
            volm_min    = st.slider("Avg Vol (Cr) >",    0.0, 10.0, 1.0, 0.1)
            circuit_max = st.slider("Circuit hits/yr <", 1, 100, 20, 1)
            circuit5    = st.slider("5% circuit 3M ≤",  0, 30, 10, 1)
        with fc3:
            use_ath   = st.checkbox("Within 25% of ATH",  value=True)
            close_min = st.slider("Min CMP ₹",            0.0, 500.0, 30.0, 5.0)

    filter_params = {
        "use_dma200": use_dma200, "use_roc12": use_roc12, "use_roc_cap": use_roc_cap,
        "volm_cr_min": volm_min, "circuit_max": circuit_max, "circuit5_max": circuit5,
        "use_away_ath": use_ath, "close_min": close_min,
    }

    U          = st.session_state.universe
    api_source = st.session_state.data_source
    end_date   = st.session_state.lookback_date

    col_run, col_info = st.columns([1, 2])
    with col_run:
        run_clicked = st.button("▶ Start Data Download", type="primary", use_container_width=True)
    with col_info:
        n_loaded = len(st.session_state.symbols) if st.session_state.symbols else "—"
        st.markdown(f"""<div style="font-size:12px;color:#64748b;padding-top:8px;">
        🌍 Universe: <b>{U}</b> &nbsp;|&nbsp;
        📊 Symbols: <b>{n_loaded}</b> &nbsp;|&nbsp;
        📅 End: <b>{end_date.strftime('%d-%m-%Y')}</b><br>
        🔢 Method: <b>{st.session_state.ranking_method}</b> &nbsp;|&nbsp;
        📡 Source: <b>{api_source}</b>
        </div>""", unsafe_allow_html=True)

    if run_clicked:
        # ── Load symbols ──────────────────────────────────────
        if st.session_state.symbols is None or st.session_state.universe != U:
            with st.spinner(f"Loading {U} symbols…"):
                try:
                    if U == "AllNSE":
                        url = f"{GITHUB_BASE}/NSE_EQ_ALL.csv"
                        df_sym = pd.read_csv(url)
                        df_sym['Yahoo_Symbol'] = df_sym['Symbol'].astype(str).str.strip() + '.NS'
                        syms_ns = df_sym['Yahoo_Symbol'].tolist()
                    else:
                        syms_ns = load_symbols_from_github(U)
                    st.session_state.symbols = syms_ns
                except Exception as e:
                    st.error(f"Symbol list load failed: {e}"); st.stop()

        symbols = st.session_state.symbols
        CHUNK   = 50 if api_source == "Upstox" else (15 if U == "AllNSE" else 50)
        st.write(f"Chunk size: **{CHUNK}** | Symbols: **{len(symbols):,}**")

        dates    = build_dates(end_date)
        prog_bar = st.progress(0)
        status_tx = st.empty()

        # Use data_service if available AND source is not YFinance or is non-YF
        _use_ds = _DS_AVAILABLE and api_source in ("Upstox", "Angel One")
        try:
            if _use_ds:
                close, high, volume, failed_symbols = fetch_data(
                    api_source   = api_source,
                    symbols      = symbols,
                    start_date   = dates['startDate'],
                    end_date     = dates['endDate'],
                    chunk_size   = CHUNK,
                    progress_bar = prog_bar,
                    status_text  = status_tx,
                )
            else:
                # YFinance inline fallback (always works)
                close, high, volume, failed_symbols = _fetch_yfinance_inline(
                    symbols, dates['startDate'], dates['endDate'],
                    prog_bar, status_tx, chunk_size=CHUNK
                )
        except Exception as e:
            st.error(f"Data fetch error: {e}"); st.stop()

        if close is None or close.empty:
            st.error("❌ Data fetch hua nahi. Internet / token check karo."); st.stop()

        # ── Failed symbols ────────────────────────────────────
        volume12M_check = volume.loc[dates['date12M']:].copy() if not volume.empty else pd.DataFrame()
        median_volume   = volume12M_check.median() if not volume12M_check.empty else pd.Series()
        failed_blank    = median_volume[median_volume.isna()].index.tolist()
        failed_blank    = [t.replace('.NS','') for t in failed_blank]
        st.session_state.failed_blank = failed_blank

        if failed_blank:
            st.warning(f"⚠ {len(failed_blank)} stocks failed to download (blank volume):")
            st.dataframe(pd.DataFrame({'S.No.': range(1, len(failed_blank)+1),
                                       'Failed Stocks': failed_blank}).set_index('S.No.'),
                         use_container_width=False)
        else:
            st.success("✅ All stocks downloaded successfully!")

        # ── Calculate metrics ─────────────────────────────────
        status_tx.markdown("⏳ **Calculating momentum metrics...**")
        prog_bar.progress(0.92)
        try:
            dfStats   = build_dfStats(close, high, volume, dates, st.session_state.ranking_method)
            dfFiltered = apply_filters(dfStats.copy(), filter_params)
            st.session_state.dfStats    = dfStats
            st.session_state.dfFiltered = dfFiltered
            st.session_state.screener_done = True
            prog_bar.progress(1.0)
            status_tx.markdown("✅ **Screener complete!**")
        except Exception as e:
            st.error(f"Calculation error: {e}"); st.stop()

    # ── Display results ───────────────────────────────────────
    if st.session_state.screener_done and st.session_state.dfFiltered is not None:
        dfF = st.session_state.dfFiltered
        dfU = st.session_state.dfStats
        n_f = len(dfF); n_u = len(dfU) if dfU is not None else 0
        top_n = st.session_state.top_n_rank
        rank_col = st.session_state.ranking_method

        st.markdown(f"""<div class="metric-row">
            {metric_card("Total Screened", f"{n_u:,}")}
            {metric_card("Passed Filters", f"{n_f:,}", "green")}
            {metric_card("Top-N Universe", f"Top {top_n}", "blue")}
            {metric_card("End Date", st.session_state.lookback_date.strftime('%d %b %Y'))}
        </div>""", unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["✅ Filtered (Top Ranked)", "📊 All Stocks (Unfiltered)"])
        with tab1:
            top_view = dfF.head(top_n).reset_index()
            dcols = ["Rank","Ticker","Close",rank_col,"roc12M","roc6M","roc3M","vol12M","volm_cr","AWAY_ATH","circuit","dma200d"]
            dcols = [c for c in dcols if c in top_view.columns]
            st.dataframe(top_view[dcols].style.format(precision=2),
                         use_container_width=True, height=440)
        with tab2:
            if dfU is not None:
                st.dataframe(dfU.reset_index().head(300).style.format(precision=2),
                             use_container_width=True, height=440)

        st.divider()
        if st.button("▶ Next: Plan Rebalance →", type="primary"):
            st.session_state.current_step = 3; st.rerun()

    elif not st.session_state.screener_done:
        st.info("⬆️ Upar se 'Start Data Download' click karo.")

# ═══════════════════════════════════════════════════════════════
# STEP 3 — PLAN REBALANCE
# ═══════════════════════════════════════════════════════════════
elif st.session_state.current_step == 3:
    st.markdown('<div class="section-hdr">Step 3 — Plan Rebalance</div>', unsafe_allow_html=True)

    if not st.session_state.screener_done:
        st.warning("⚠️ Pehle Step 2 mein screener run karo.")
        if st.button("← Step 2 par jao"): st.session_state.current_step = 2; st.rerun()
        st.stop()

    # ── Portfolio source ──────────────────────────────────────
    port_source = st.radio("Portfolio data source",
                           ["📊 Google Sheet (auto)", "📂 CSV manually upload"],
                           horizontal=True)

    if "📊" in port_source:
        col_load, _ = st.columns([1, 2])
        with col_load:
            if st.button("🔄 Fetch from Google Sheet", type="primary"):
                with st.spinner("Fetching portfolio..."):
                    try:
                        pdf = pd.read_csv(PORTFOLIO_CSV_URL)
                        if 'Current Portfolio' in pdf.columns:
                            st.session_state.reb_portfolio = [
                                str(x).strip().upper() for x in pdf['Current Portfolio'].dropna()
                                if str(x).strip() and str(x).strip().lower() not in ('nan','current portfolio','')
                            ]
                            st.success(f"✅ Portfolio loaded: **{len(st.session_state.reb_portfolio)}** stocks")
                        else:
                            st.error("Column 'Current Portfolio' not found in sheet.")
                    except Exception as e:
                        st.error(f"Sheet fetch failed: {e}")
    else:
        up_reb = st.file_uploader("📂 Upload Portfolio CSV", type=["csv"], key="reb_csv")
        if up_reb:
            try:
                df_reb = pd.read_csv(up_reb)
                df_reb.columns = [c.strip() for c in df_reb.columns]
                col_b = df_reb.columns[1] if len(df_reb.columns) > 1 else df_reb.columns[0]
                st.session_state.reb_portfolio = [
                    str(x).strip().upper() for x in df_reb[col_b].dropna()
                    if str(x).strip() and len(str(x).strip()) > 1
                ]
                st.success(f"✅ Portfolio loaded: {len(st.session_state.reb_portfolio)} stocks")
            except Exception as e:
                st.error(f"CSV parse error: {e}")

    # ── Manual override ───────────────────────────────────────
    with st.expander("✏️ Manual Edit (comma-separated)", expanded=False):
        port_text = st.text_area("Current Portfolio",
                                  value=", ".join(st.session_state.reb_portfolio or []), height=100)
        if st.button("Apply Manual Edit"):
            st.session_state.reb_portfolio = [
                s.strip().upper() for s in port_text.split(",") if s.strip()
            ]
            st.success("Updated!")

    portfolio = st.session_state.reb_portfolio or []

    # ── Compute rebalance ─────────────────────────────────────
    if portfolio and st.session_state.dfFiltered is not None:
        dfFiltered      = st.session_state.dfFiltered
        dfStats         = st.session_state.dfStats
        top_n           = st.session_state.top_n_rank
        rank_threshold  = top_n

        top_rank_tickers = dfFiltered.reset_index()
        top_rank_tickers = top_rank_tickers[top_rank_tickers['Rank'] <= rank_threshold]['Ticker']

        current_portfolio_tickers = pd.Series(portfolio)
        entry_stocks = top_rank_tickers[~top_rank_tickers.isin(current_portfolio_tickers)]
        exit_stocks  = current_portfolio_tickers[~current_portfolio_tickers.isin(top_rank_tickers)]
        hold_stocks  = current_portfolio_tickers[current_portfolio_tickers.isin(top_rank_tickers)]

        num_sells = len(exit_stocks)
        entry_stocks = entry_stocks.head(num_sells)

        if len(entry_stocks) < num_sells:
            entry_stocks = pd.concat([
                entry_stocks,
                pd.Series([None] * (num_sells - len(entry_stocks)))
            ])

        # ── Reasons for exit (v10 logic) ──────────────────────
        reasons_for_exit = []
        for ticker in exit_stocks:
            if pd.isna(ticker) or ticker == "":
                reasons_for_exit.append(""); continue
            reasons    = []
            stock_data = dfStats[dfStats['Ticker'] == ticker] if dfStats is not None else pd.DataFrame()
            if len(stock_data) > 0:
                if stock_data.index[0] > rank_threshold:          reasons.append(f"Rank > {rank_threshold}")
                if stock_data['volm_cr'].values[0] <= 1:           reasons.append("Volume ≤ 1 Cr")
                if stock_data['Close'].values[0] <= stock_data['dma200d'].values[0]:
                                                                   reasons.append("Close ≤ 200-DMA")
                if stock_data['roc12M'].values[0] <= 5.5:          reasons.append("12M ROC ≤ 5.5%")
                if stock_data['circuit'].values[0] >= 20:          reasons.append("Circuit ≥ 20")
                if stock_data['AWAY_ATH'].values[0] <= -25:        reasons.append("Away ATH ≤ -25%")
                if stock_data['roc12M'].values[0] >= 1000:         reasons.append("12M ROC ≥ 1000%")
                if stock_data['Close'].values[0] <= 30:            reasons.append("Close ≤ ₹30")
                if stock_data['circuit5'].values[0] > 10:          reasons.append("5% Circuit > 10")
            else:
                reasons.append("Not in selected universe")
            reasons_for_exit.append(", ".join(reasons) if reasons else "Rank dropped")

        reasons_for_exit.extend([""] * (len(entry_stocks) - len(reasons_for_exit)))

        rebalance_table = pd.DataFrame({
            'S.No.':           range(1, num_sells + 1),
            'Sell Stocks':     exit_stocks.tolist(),
            'Buy Stocks':      entry_stocks.tolist(),
            'Reason for Exit': reasons_for_exit,
        })
        rebalance_table = rebalance_table[
            ~(rebalance_table['Sell Stocks'].isna() & rebalance_table['Buy Stocks'].isna())
        ]
        rebalance_table.set_index('S.No.', inplace=True)
        st.session_state.sell_list = exit_stocks.dropna().tolist()
        st.session_state.buy_list  = entry_stocks.dropna().tolist()
        st.session_state.rebalance_table = rebalance_table
        st.session_state.rebalance_done  = True

        # ── Summary strip ──────────────────────────────────────
        st.markdown(f"""<div class="reb-strip">
          <div class="reb-stat"><div class="label">Portfolio</div><div class="val b">{len(portfolio)}</div></div>
          <div class="reb-stat"><div class="label">Top-{rank_threshold} Screener</div><div class="val b">{len(top_rank_tickers)}</div></div>
          <div class="reb-stat"><div class="label">SELL (Exit)</div><div class="val r">{len(exit_stocks)}</div></div>
          <div class="reb-stat"><div class="label">BUY (Entry)</div><div class="val g">{len(entry_stocks.dropna())}</div></div>
          <div class="reb-stat"><div class="label">HOLD</div><div class="val p">{len(hold_stocks)}</div></div>
        </div>""", unsafe_allow_html=True)

        # ── Sell / Buy / Hold columns ──────────────────────────
        col_sell, col_buy, col_hold = st.columns(3)
        with col_sell:
            st.markdown("#### 🔴 SELL List")
            sell_list = exit_stocks.dropna().tolist()
            if sell_list:
                chips = " ".join([f'<span class="chip chip-sell">{s}</span>' for s in sell_list])
                st.markdown(chips, unsafe_allow_html=True)
                cmp_map = {}
                if dfStats is not None:
                    cmp_map = dict(zip(dfStats['Ticker'], dfStats['Close']))
                # Also try dfFiltered for CMP (in case stock is in filtered but not dfStats)
                if dfFiltered is not None:
                    for t, c in zip(dfFiltered.reset_index()['Ticker'], dfFiltered.reset_index()['Close']):
                        if t not in cmp_map:
                            cmp_map[t] = c
                sell_df = pd.DataFrame({
                    "Stock": sell_list,
                    "CMP ₹": [
                        round(cmp_map[s], 2) if s in cmp_map and cmp_map[s] > 0
                        else "N/A *"
                        for s in sell_list
                    ],
                    "Reason": reasons_for_exit[:len(sell_list)]
                })
                st.dataframe(sell_df, hide_index=True, use_container_width=True)
                missing_cmp = [s for s in sell_list if s not in cmp_map or cmp_map.get(s, 0) == 0]
                if missing_cmp:
                    st.caption(
                        f"* {', '.join(missing_cmp)} — CMP unavailable "
                        f"(stock selected universe ({st.session_state.universe}) mein nahi hai). "
                        "Broker app se manually CMP check karo."
                    )
            else:
                st.success("Koi sell nahi hai!")

        with col_buy:
            st.markdown("#### 🟢 BUY List (New Entry)")
            buy_list = entry_stocks.dropna().tolist()
            if buy_list:
                chips = " ".join([f'<span class="chip chip-buy">{s}</span>' for s in buy_list])
                st.markdown(chips, unsafe_allow_html=True)
                rank_map = dict(zip(dfFiltered.reset_index()['Ticker'], dfFiltered.reset_index()['Rank']))
                cmp_map2 = {}
                if dfStats is not None:
                    cmp_map2 = dict(zip(dfStats['Ticker'], dfStats['Close']))
                buy_df = pd.DataFrame({
                    "Stock":        buy_list,
                    "Screener Rank":[rank_map.get(s, "—") for s in buy_list],
                    "CMP ₹":        [round(cmp_map2.get(s, 0), 2) for s in buy_list],
                })
                st.dataframe(buy_df, hide_index=True, use_container_width=True)
            else:
                st.info("Koi buy nahi hai.")

        with col_hold:
            st.markdown("#### 🔵 HOLD (Retain)")
            if not hold_stocks.empty:
                chips = " ".join([f'<span class="chip chip-hold">{s}</span>' for s in hold_stocks.tolist()])
                st.markdown(chips, unsafe_allow_html=True)

        # ── Rebalance table ────────────────────────────────────
        st.markdown('<div class="section-hdr">📋 Rebalance Table (Sell → Buy mapping)</div>', unsafe_allow_html=True)
        if not rebalance_table.empty:
            st.dataframe(rebalance_table, use_container_width=True)

        st.divider()

        # ══════════════════════════════════════════════════════════
        # WORKFLOW PANEL — Screener → Rebalancer → Order Calculator
        # ══════════════════════════════════════════════════════════
        st.markdown('<div class="section-hdr">🔄 Rebalancer Workflow</div>', unsafe_allow_html=True)

        # ── Step A: Copy Top-N screener list → Google Sheet "Worst Rank Held"
        sell_list_local = exit_stocks.dropna().tolist()
        buy_list_local  = entry_stocks.dropna().tolist()

        cmp_map3 = {}
        if dfStats is not None:
            cmp_map3 = dict(zip(dfStats['Ticker'], dfStats['Close']))

        # Top-N screener tickers — Worst Rank Held column ke liye
        # Sirf wahi stocks jo filter pass kiye AND rank <= top_n_rank
        # (Excel "Filtered Stocks" sheet ke same 48 stocks)
        _df_sorted = dfFiltered.reset_index()
        if 'Rank' in _df_sorted.columns:
            _df_sorted = _df_sorted.sort_values('Rank', ascending=True)
            _top_filtered = _df_sorted[_df_sorted['Rank'] <= st.session_state.top_n_rank]
        else:
            _top_filtered = _df_sorted.head(st.session_state.top_n_rank)
        top_n_tickers = _top_filtered["Ticker"].tolist()
        worst_rank_text = "\n".join(top_n_tickers) if top_n_tickers else "(no data)"

        st.markdown("""
        <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;
                    padding:14px 18px;margin:8px 0;font-size:13px;line-height:2;">
        <b>📋 Workflow Steps:</b><br>
        <b>1.</b> Neeche <b>Top-N Screener list</b> copy karo → Google Sheet ke <b>"Worst Rank Held"</b> column mein paste karo
          <span style="color:#64748b;font-size:12px;">(ye list rebalancer ko batati hai ki kaun good rank mein hai)</span><br>
        <b>2.</b> <b>"Open Portfolio Rebalancer"</b> button dabao → Sell stocks select karo → actual sell value note karo<br>
        <b>3.</b> Woh sell value neeche <b>"Sell Value"</b> field mein enter karo → Buy orders auto-calculate honge
        </div>
        """, unsafe_allow_html=True)

        wa1, wa2 = st.columns([1, 1])
        with wa1:
            n_top = len(top_n_tickers)
            st.markdown(f"**📋 Top-{st.session_state.top_n_rank} Screener List — Google Sheet 'Worst Rank Held' column mein paste karo:**")
            st.caption(f"✅ {n_top} filtered & ranked stocks | Rank 1 se Rank {n_top} tak")
            st.text_area(
                "Top-N list — Google Sheet Worst Rank Held column mein paste karo",
                value=worst_rank_text,
                height=min(160, max(80, len(top_n_tickers) * 6 + 60)),
                key="sell_copy_area",
                label_visibility="collapsed",
                help="Yeh Top-N screener stocks Google Sheet ke Worst Rank Held column mein paste karo"
            )
            # Clipboard copy — uses execCommand fallback for Streamlit iframe sandbox
            import streamlit.components.v1 as _components
            _safe_text = worst_rank_text.replace("`", "'").replace("\\", "/")
            _copy_html = f"""
            <textarea id="cpytxt" style="position:absolute;left:-9999px;">{_safe_text}</textarea>
            <button id="cpybtn"
              onclick="
                var t=document.getElementById('cpytxt');
                t.select(); t.setSelectionRange(0,99999);
                var ok=false;
                try{{ok=document.execCommand('copy');}}catch(e){{}}
                if(!ok && navigator.clipboard){{
                  navigator.clipboard.writeText(t.value).then(function(){{
                    document.getElementById('cpybtn').innerHTML='✅ Copied!';
                    document.getElementById('cpybtn').style.background='#16a34a';
                  }});
                }} else if(ok) {{
                  document.getElementById('cpybtn').innerHTML='✅ Copied!';
                  document.getElementById('cpybtn').style.background='#16a34a';
                }} else {{
                  alert('Manually select text above aur Ctrl+C / Cmd+C dabao');
                }}
              "
              style="background:#2563eb;color:white;border:none;padding:8px 20px;
                     border-radius:6px;font-weight:700;cursor:pointer;font-size:13px;
                     margin-top:4px;">
              📋 Copy to Clipboard
            </button>
            """
            _components.html(_copy_html, height=50)

        with wa2:
            st.markdown("**⚖️ Portfolio Rebalancer:**")
            st.markdown(f"""
            <a href="{APPS_SCRIPT_URL}" target="_blank">
            <div style="background:#1a237e;color:white;padding:12px 20px;border-radius:8px;
                        text-align:center;font-weight:700;font-size:14px;margin:4px 0;
                        cursor:pointer;text-decoration:none;">
              ⚖️ Open Portfolio Rebalancer
            </div></a>
            <div style="font-size:11px;color:#64748b;margin-top:6px;line-height:1.6;">
              Wahan se sell karke <b>actual sell value</b> note karo.<br>
              Phir neeche woh value enter karo.
            </div>
            """, unsafe_allow_html=True)

        st.divider()

        # ── Step B: Order Calculator ───────────────────────────────
        st.markdown('<div class="section-hdr">⚡ Order Calculator</div>', unsafe_allow_html=True)

        qr1, qr2, qr3, qr4 = st.columns(4)
        with qr1:
            capital_add = st.number_input(
                "💰 Capital Addition ₹", min_value=0, value=0, step=5000, key="qr_cap"
            )
        with qr2:
            brokerage = st.number_input(
                "🏦 Brokerage/Stock ₹", min_value=0, value=0, step=10, key="qr_brk"
            )
        with qr3:
            sell_val_input = st.number_input(
                "💸 Sell Value ₹ (Rebalancer se enter karo)",
                min_value=0, value=0, step=1000, key="qr_sell",
                help="Portfolio Rebalancer mein jo actual sell value mili, woh yahaan enter karo"
            )

        sell_brk  = len(sell_list_local) * brokerage
        buy_brk   = len(buy_list_local)  * brokerage
        net_pool  = sell_val_input + capital_add - sell_brk - buy_brk
        per_stock = net_pool / len(buy_list_local) if buy_list_local else 0

        with qr4:
            st.markdown(f"""<div class="metric-card">
              <div class="metric-label">Net Pool / Stock</div>
              <div class="metric-value green">{fmt_inr(per_stock)}</div>
            </div>""", unsafe_allow_html=True)

        # Summary strip
        st.markdown(f"""<div class="reb-strip">
          <div class="reb-stat"><div class="label">Sell Value</div><div class="val b">₹{sell_val_input:,.0f}</div></div>
          <div class="reb-stat"><div class="label">+ Capital</div><div class="val g">₹{capital_add:,.0f}</div></div>
          <div class="reb-stat"><div class="label">- Sell Brok</div><div class="val r">₹{sell_brk:,.0f}</div></div>
          <div class="reb-stat"><div class="label">- Buy Brok</div><div class="val r">₹{buy_brk:,.0f}</div></div>
          <div class="reb-stat"><div class="label">Net Pool</div><div class="val g">₹{net_pool:,.0f}</div></div>
          <div class="reb-stat"><div class="label">Per Stock</div><div class="val g">{fmt_inr(per_stock)}</div></div>
        </div>""", unsafe_allow_html=True)

        if sell_val_input == 0 and not capital_add:
            st.info("💡 Sell Value enter karo (Portfolio Rebalancer se) → Buy orders auto-calculate honge.")

        # ── Buy orders table ────────────────────────────────────────
        if buy_list_local and per_stock > 0:
            st.markdown('<div class="section-hdr">📋 Buy Orders (Estimated)</div>', unsafe_allow_html=True)
            orders = []
            total_invested = 0
            for i, stock in enumerate(buy_list_local, 1):
                cmp = cmp_map3.get(stock, 0)
                if cmp > 0:
                    qty = int(per_stock / cmp)
                    val = qty * cmp
                    total_invested += val
                    orders.append({
                        "#":           i,
                        "Stock":       stock,
                        "CMP ₹":       round(cmp, 2),
                        "Gross Alloc": round(per_stock + brokerage),
                        "Brok ₹":      brokerage,
                        "Net Alloc":   round(per_stock),
                        "Qty":         qty,
                        "Value ₹":     round(val),
                    })

            if orders:
                orders_df = pd.DataFrame(orders)
                st.dataframe(
                    orders_df.style.format({
                        "CMP ₹": "{:.2f}",
                        "Gross Alloc": "{:,.0f}", "Net Alloc": "{:,.0f}", "Value ₹": "{:,.0f}"
                    }),
                    use_container_width=True, hide_index=True, height=300
                )
                leftover = net_pool - total_invested
                st.markdown(f"""<div class="reb-strip">
                  <div class="reb-stat"><div class="label">Total Invested</div><div class="val g">₹{total_invested:,.0f}</div></div>
                  <div class="reb-stat"><div class="label">Leftover</div><div class="val p">₹{leftover:,.0f}</div></div>
                  <div class="reb-stat"><div class="label">Buy Orders</div><div class="val b">{len(orders)}</div></div>
                </div>""", unsafe_allow_html=True)

        st.divider()
        if st.button("▶ Next: Apply & Export →", type="primary"):
            st.session_state.current_step = 4; st.rerun()

    elif not portfolio:
        st.info("⬆️ Upar se portfolio data load karo (Google Sheet ya CSV).")


# ═══════════════════════════════════════════════════════════════
# STEP 4 — APPLY & EXPORT
# ═══════════════════════════════════════════════════════════════
elif st.session_state.current_step == 4:
    st.markdown('<div class="section-hdr">Step 4 — Apply & Export</div>', unsafe_allow_html=True)

    sell        = st.session_state.sell_list or []
    buy         = st.session_state.buy_list  or []
    portfolio   = st.session_state.reb_portfolio or []
    dfStats     = st.session_state.dfStats
    dfFiltered  = st.session_state.dfFiltered
    reb_table   = st.session_state.rebalance_table
    failed_blank= st.session_state.failed_blank or []
    U           = st.session_state.universe
    top_n       = st.session_state.top_n_rank
    rank_method = st.session_state.ranking_method
    api_source  = st.session_state.data_source
    end_date    = st.session_state.lookback_date

    # ── Summary ───────────────────────────────────────────────
    st.markdown(f"""<div class="reb-strip">
      <div class="reb-stat"><div class="label">Exits (SELL)</div><div class="val r">{len(sell)}</div></div>
      <div class="reb-stat"><div class="label">New Entries (BUY)</div><div class="val g">{len(buy)}</div></div>
      <div class="reb-stat"><div class="label">Retained (HOLD)</div><div class="val p">{len(portfolio) - len(sell)}</div></div>
      <div class="reb-stat"><div class="label">New Portfolio Size</div><div class="val b">{len(portfolio) - len(sell) + len(buy)}</div></div>
    </div>""", unsafe_allow_html=True)

    # ── v10-format Excel Export ────────────────────────────────
    if dfFiltered is not None and dfStats is not None:
        st.markdown('<div class="section-hdr">💾 Excel Export (v10 Format — 4 Sheets)</div>', unsafe_allow_html=True)

        # ── Failed Downloads DF ──────────────────────────────
        if failed_blank:
            df_failed = pd.DataFrame({
                'S.No.':        range(1, len(failed_blank)+1),
                'Failed Stock': failed_blank
            }).set_index('S.No.')
        else:
            df_failed = pd.DataFrame(columns=['Failed Stock'])
            df_failed.index.name = 'S.No.'

        # ── Rebalance Table ──────────────────────────────────
        if reb_table is None or reb_table.empty:
            reb_table = pd.DataFrame(columns=['Sell Stocks','Buy Stocks','Reason for Exit'])
            reb_table.index.name = 'S.No.'

        filtered = dfFiltered.copy()

        excel_file = f"{end_date.strftime('%Y-%m-%d')}_{U}_{rank_method}_{api_source}_lookback.xlsx"

        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            dfStats.to_excel(   writer, sheet_name="Unfiltered Stocks",     index=True)
            filtered.to_excel(  writer, sheet_name="Filtered Stocks",       index=True)
            df_failed.to_excel( writer, sheet_name="Failed Downloads",      index=True)
            reb_table.to_excel( writer, sheet_name="Portfolio Rebalancing", index=True)

        # Apply v10 formatting
        try:
            format_excel_unfiltered(excel_file, U, top_n)
            format_excel_filtered(excel_file, U, top_n)
            format_simple_sheet(excel_file, "Failed Downloads")
            format_simple_sheet(excel_file, "Portfolio Rebalancing")
        except Exception as e:
            st.warning(f"Excel formatting partial error (file still usable): {e}")

        with open(excel_file, "rb") as f:
            st.download_button(
                label     = "📥 Download Excel (4 Sheets)",
                data      = f.read(),
                file_name = excel_file,
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type      = "primary",
            )

        st.success(f"✅ Excel ready: `{excel_file}`")
        st.caption("Sheets: Unfiltered Stocks | Filtered Stocks | Failed Downloads | Portfolio Rebalancing")

    # ── Apps Script / Rebalance Sheet links ───────────────────
    st.markdown('<div class="section-hdr">📊 Apps Script Workflow — Quick Links</div>', unsafe_allow_html=True)

    col_links = st.columns(3)
    with col_links[0]:
        st.markdown(f"""
        <a href="{APPS_SCRIPT_URL}" target="_blank">
        <div style="background:#1a237e;color:white;padding:10px 16px;border-radius:8px;
                    text-align:center;font-weight:700;font-size:13px;">
        ⚖️ Portfolio Rebalancer
        </div></a>
        """, unsafe_allow_html=True)
    with col_links[1]:
        st.markdown("""
        <a href="https://docs.google.com/spreadsheets/d/1xb8xoW91HWeXBW8Zd99TobULSgwxcvfPaaYPlMLZmHI" target="_blank">
        <div style="background:#2e7d32;color:white;padding:10px 16px;border-radius:8px;
                    text-align:center;font-weight:700;font-size:13px;">
        📊 Rebalance Sheet
        </div></a>
        """, unsafe_allow_html=True)
    with col_links[2]:
        st.markdown("""
        <a href="https://prayan2702.github.io/momn-dashboard/" target="_blank">
        <div style="background:#7c3aed;color:white;padding:10px 16px;border-radius:8px;
                    text-align:center;font-weight:700;font-size:13px;">
        📈 Portfolio Dashboard
        </div></a>
        """, unsafe_allow_html=True)

    # ── Rebalancing table on screen ───────────────────────────
    if reb_table is not None and not reb_table.empty:
        st.markdown('<div class="section-hdr">📋 Portfolio Rebalancing</div>', unsafe_allow_html=True)
        st.dataframe(reb_table, use_container_width=True)

    # ── Next Steps Reminder ───────────────────────────────────
    st.markdown("### 📌 Next Steps — Rebalance Ke Baad")
    st.markdown("**🏁 Trades execute hone ke baad yeh karo:**")

    col_ns1, col_ns2 = st.columns(2)
    with col_ns1:
        st.success("**1. 📊 Rebalance Sheet update karo**\n\nPortfolio sheet mein sell wale stocks hato, buy wale add karo.")
        st.info("**3. 📈 XIRR / Dashboard refresh karo**\n\nPortfolio value update hone ke baad XIRR recalculate hogi.")
    with col_ns2:
        st.warning("**2. 💾 Monthly Backup lo**\n\nGoogle Sheet → File → Make a copy → rename: `Portfolio_Backup_MMYYYY`")
        st.markdown("""
        <div style="background:#f3e8ff;border:1px solid #d8b4fe;border-radius:8px;padding:12px 16px;font-size:13px;">
        🗓️ <b>Next Rebalance Date note karo</b><br>
        March-end ya September-end (jo bhi agle aaye)
        </div>
        """, unsafe_allow_html=True)

    st.divider()
    col_r1, col_r2 = st.columns(2)
    with col_r1:
        if st.button("🔄 New Month — Restart from Step 1", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()
    with col_r2:
        if st.button("← Step 3 — Edit Rebalance"):
            st.session_state.current_step = 3; st.rerun()
